import csv
import io
import json
import os
import re
import textwrap
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas


APP_TITLE = "Inspector Equipment – Need List"
DATA_DIR = "data"
ITEMS_PATH = os.path.join(DATA_DIR, "items.json")
SUBMISSIONS_PATH = os.path.join(DATA_DIR, "submissions.csv")
REQUEST_HISTORY_PATH = os.path.join(DATA_DIR, "request_history.csv")
ADMIN_SETTINGS_PATH = os.path.join(DATA_DIR, "admin_settings.json")


@dataclass
class Item:
    label: str
    value_field: str = "none"   # "none" | "text" | "number" | "choice"
    choices: Optional[List[str]] = None


# ---------- Storage helpers ----------
def ensure_data_dir() -> None:
    os.makedirs(DATA_DIR, exist_ok=True)


def load_items() -> List[Item]:
    ensure_data_dir()

    if not os.path.exists(ITEMS_PATH):
        with open(ITEMS_PATH, "w", encoding="utf-8") as f:
            json.dump({"items": []}, f, indent=2)

    with open(ITEMS_PATH, "r", encoding="utf-8") as f:
        raw = json.load(f)

    items: List[Item] = []
    for it in raw.get("items", []):
        label = str(it.get("label", "")).strip()
        value_field = str(it.get("value_field", "none")).strip().lower()
        choices = it.get("choices")

        if label:
            items.append(
                Item(
                    label=label,
                    value_field=value_field if value_field in {"none", "text", "number", "choice"} else "none",
                    choices=choices if isinstance(choices, list) else None,
                )
            )

    return items


def save_items(items: List[Item]) -> None:
    ensure_data_dir()
    payload = {
        "items": [
            {
                "label": i.label,
                "value_field": i.value_field,
                **({"choices": i.choices} if i.value_field == "choice" and i.choices else {}),
            }
            for i in items
        ]
    }
    with open(ITEMS_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)


def ensure_submissions_file() -> None:
    ensure_data_dir()
    if not os.path.exists(SUBMISSIONS_PATH):
        with open(SUBMISSIONS_PATH, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["timestamp_utc", "inspector_name", "needed_json", "comment"])


def append_submission(inspector_name: str, needed: List[Dict[str, Any]], comment: str) -> None:
    ensure_submissions_file()
    ts = datetime.now(timezone.utc).isoformat()
    with open(SUBMISSIONS_PATH, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([ts, inspector_name.strip(), json.dumps(needed, ensure_ascii=False), comment.strip()])


def ensure_request_history_file() -> None:
    ensure_data_dir()
    if not os.path.exists(REQUEST_HISTORY_PATH):
        with open(REQUEST_HISTORY_PATH, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(
                [
                    "timestamp_utc",
                    "date_requested",
                    "time_requested",
                    "year_month",
                    "inspector_name",
                    "items_json",
                    "items_display",
                    "truck_model_year",
                    "truck_unit_number",
                    "comment",
                    "pdf_filename",
                ]
            )


def append_request_history(
    inspector_name: str,
    item_entries: List[Dict[str, Any]],
    comment: str,
    truck_model_year: Optional[int],
    truck_unit_number: Optional[int],
    pdf_filename: str,
) -> None:
    ensure_request_history_file()

    now_local = datetime.now()
    timestamp_utc = datetime.now(timezone.utc).isoformat()
    date_requested = now_local.strftime("%Y-%m-%d")
    time_requested = now_local.strftime("%I:%M %p").lstrip("0")
    year_month = now_local.strftime("%Y-%m")

    display_parts = []
    for entry in item_entries:
        quantity = int(entry.get("quantity", 0))
        item_name = str(entry.get("item", "")).strip()
        value = entry.get("value", "")
        item_text = item_name
        if value not in (None, "", "0"):
            item_text = f"{item_name} ({value})"
        display_parts.append(f"{quantity} x {item_text}")

    items_display = ", ".join(display_parts)

    with open(REQUEST_HISTORY_PATH, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(
            [
                timestamp_utc,
                date_requested,
                time_requested,
                year_month,
                inspector_name.strip(),
                json.dumps(item_entries, ensure_ascii=False),
                items_display,
                "" if truck_model_year in (None, 0) else int(truck_model_year),
                "" if truck_unit_number in (None, 0) else int(truck_unit_number),
                comment.strip(),
                pdf_filename,
            ]
        )


def load_request_history_rows() -> List[Dict[str, str]]:
    ensure_request_history_file()
    rows: List[Dict[str, str]] = []
    with open(REQUEST_HISTORY_PATH, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append({k: (v if v is not None else "") for k, v in row.items()})
    return rows


def download_csv_button() -> None:
    if not os.path.exists(SUBMISSIONS_PATH):
        return

    with open(SUBMISSIONS_PATH, "rb") as f:
        st.download_button(
            label="Download submissions CSV",
            data=f.read(),
            file_name="submissions.csv",
            mime="text/csv",
            use_container_width=True,
        )


def download_request_history_csv_button() -> None:
    if not os.path.exists(REQUEST_HISTORY_PATH):
        return

    with open(REQUEST_HISTORY_PATH, "rb") as f:
        st.download_button(
            label="Download request history CSV",
            data=f.read(),
            file_name="request_history.csv",
            mime="text/csv",
            use_container_width=True,
        )


# ---------- Admin settings ----------
def load_admin_settings() -> Dict[str, Any]:
    ensure_data_dir()

    defaults = {
        "submit_popup_enabled": False,
        "submit_popup_message": "",
    }

    if not os.path.exists(ADMIN_SETTINGS_PATH):
        with open(ADMIN_SETTINGS_PATH, "w", encoding="utf-8") as f:
            json.dump(defaults, f, indent=2)
        return defaults

    try:
        with open(ADMIN_SETTINGS_PATH, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except Exception:
        raw = {}

    settings = {
        "submit_popup_enabled": bool(raw.get("submit_popup_enabled", False)),
        "submit_popup_message": str(raw.get("submit_popup_message", "")).strip(),
    }
    return settings


def save_admin_settings(settings: Dict[str, Any]) -> None:
    ensure_data_dir()
    payload = {
        "submit_popup_enabled": bool(settings.get("submit_popup_enabled", False)),
        "submit_popup_message": str(settings.get("submit_popup_message", "")).strip(),
    }
    with open(ADMIN_SETTINGS_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)


# ---------- Utility helpers ----------
def safe_filename(value: str) -> str:
    cleaned = "".join(ch for ch in value if ch.isalnum() or ch in (" ", "_", "-")).strip()
    return cleaned or "Inspector"


def wrap_text_lines(text: str, width: int = 90) -> List[str]:
    if not text:
        return [""]
    lines: List[str] = []
    for paragraph in str(text).splitlines() or [""]:
        wrapped = textwrap.wrap(paragraph, width=width) if paragraph else [""]
        lines.extend(wrapped)
    return lines or [""]


def item_key_from_label(label: str) -> str:
    base = re.sub(r"[^a-zA-Z0-9]+", "_", label.strip().lower()).strip("_")
    return base or "item"


def is_truck_field(label: str) -> bool:
    l = label.strip().lower()
    return l in {"truck model year", "truck unit number"}


def value_is_meaningful(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    if isinstance(value, (int, float)):
        return value != 0
    return True


def parse_items_json(items_json: str) -> List[Dict[str, Any]]:
    if not items_json.strip():
        return []
    try:
        raw = json.loads(items_json)
        if isinstance(raw, list):
            cleaned: List[Dict[str, Any]] = []
            for entry in raw:
                if isinstance(entry, dict):
                    cleaned.append(
                        {
                            "item": str(entry.get("item", "")).strip(),
                            "quantity": int(entry.get("quantity", 0) or 0),
                            "value": entry.get("value", ""),
                        }
                    )
            return cleaned
    except Exception:
        return []
    return []


def get_available_months(rows: List[Dict[str, str]]) -> List[str]:
    months = sorted({row.get("year_month", "").strip() for row in rows if row.get("year_month", "").strip()}, reverse=True)
    return months


def compute_item_totals_from_rows(rows: List[Dict[str, str]]) -> Dict[str, int]:
    totals: Dict[str, int] = defaultdict(int)
    for row in rows:
        for entry in parse_items_json(row.get("items_json", "")):
            item_name = str(entry.get("item", "")).strip()
            quantity = int(entry.get("quantity", 0) or 0)
            if item_name and quantity > 0:
                totals[item_name] += quantity
    return dict(sorted(totals.items(), key=lambda x: x[0].lower()))


def compute_lifetime_item_totals(rows: List[Dict[str, str]]) -> Dict[str, int]:
    return compute_item_totals_from_rows(rows)


# ---------- PDF helpers ----------
def create_receipt_pdf(
    inspector_name: str,
    item_entries: List[Dict[str, Any]],
    comment: str,
    truck_model_year: Optional[int],
    truck_unit_number: Optional[int],
) -> Tuple[str, bytes]:
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    _, page_height = letter

    left_margin = 50
    top_y = page_height - 50
    y = top_y

    def new_page() -> None:
        nonlocal y
        pdf.showPage()
        y = top_y

    def draw_text(x: int, y_pos: int, text: str, font_name: str = "Helvetica", font_size: int = 11) -> None:
        pdf.setFont(font_name, font_size)
        pdf.drawString(x, y_pos, text)

    def require_space(lines_needed: int = 1, line_height: int = 18) -> None:
        nonlocal y
        if y - (lines_needed * line_height) < 60:
            new_page()

    pdf.setTitle("Equipment Request Receipt")

    request_date = datetime.now().strftime("%Y-%m-%d")
    request_time = datetime.now().strftime("%I:%M %p").lstrip("0")

    require_space(6)
    draw_text(left_margin, y, "Date Requested", "Helvetica-Bold", 12)
    draw_text(left_margin + 150, y, request_date, "Helvetica", 12)
    draw_text(left_margin + 310, y, "INSPECTOR", "Helvetica-Bold", 12)
    draw_text(left_margin + 420, y, inspector_name.strip(), "Helvetica", 12)
    y -= 28

    draw_text(left_margin, y, "Time Requested", "Helvetica-Bold", 12)
    draw_text(left_margin + 150, y, request_time, "Helvetica", 12)
    y -= 30

    draw_text(left_margin, y, "QUANTITY", "Helvetica-Bold", 12)
    draw_text(left_margin + 120, y, "EQUIPMENT ITEM", "Helvetica-Bold", 12)
    y -= 14

    pdf.line(left_margin, y, left_margin + 500, y)
    y -= 20

    if item_entries:
        for entry in item_entries:
            quantity = int(entry.get("quantity", 0) or 0)
            item_name = str(entry.get("item", "")).strip()
            value = entry.get("value", "")
            item_text = item_name if item_name else ""

            if value_is_meaningful(value):
                item_text = f"{item_text} ({value})"

            wrapped_lines = wrap_text_lines(item_text, width=58)
            require_space(len(wrapped_lines) + 1)

            draw_text(left_margin, y, str(quantity), "Helvetica", 12)
            draw_text(left_margin + 120, y, wrapped_lines[0], "Helvetica", 12)
            y -= 18

            for extra_line in wrapped_lines[1:]:
                draw_text(left_margin + 120, y, extra_line, "Helvetica", 12)
                y -= 18
    else:
        require_space(2)
        draw_text(left_margin + 120, y, "No items selected.", "Helvetica", 12)
        y -= 22

    additional_lines: List[str] = []
    if truck_model_year not in (None, 0):
        additional_lines.append(f"Truck Model Year: {int(truck_model_year)}")
    if truck_unit_number not in (None, 0):
        additional_lines.append(f"Truck Unit Number: {int(truck_unit_number)}")

    if additional_lines:
        y -= 10
        require_space(len(additional_lines) + 2)
        draw_text(left_margin, y, "ADDITIONAL DETAILS", "Helvetica-Bold", 12)
        y -= 18
        for line in additional_lines:
            draw_text(left_margin, y, line, "Helvetica", 12)
            y -= 18

    y -= 10
    require_space(4)
    draw_text(left_margin, y, "COMMENTS", "Helvetica-Bold", 12)
    y -= 18

    if comment.strip():
        for wrapped in wrap_text_lines(comment.strip(), width=88):
            require_space(1)
            draw_text(left_margin, y, wrapped, "Helvetica", 11)
            y -= 16
    else:
        draw_text(left_margin, y, "None", "Helvetica", 11)
        y -= 16

    pdf.save()
    pdf_bytes = buffer.getvalue()
    buffer.close()

    filename = f"Equipment requested by {safe_filename(inspector_name)} {request_date}.pdf"
    return filename, pdf_bytes


# ---------- Excel export helpers ----------
def build_monthly_workbook(
    selected_month: str,
    month_rows: List[Dict[str, str]],
) -> bytes:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Monthly Request Receipts"

    bold = Font(bold=True)
    thin_side = Side(style="thin", color="000000")
    bottom_border = Border(bottom=thin_side)

    ws1.column_dimensions["A"].width = 14
    ws1.column_dimensions["B"].width = 4
    ws1.column_dimensions["C"].width = 34
    ws1.column_dimensions["D"].width = 4
    ws1.column_dimensions["E"].width = 14
    ws1.column_dimensions["F"].width = 4
    ws1.column_dimensions["G"].width = 28

    row_cursor = 1

    if not month_rows:
        ws1["A1"] = "No requests found for selected month."
    else:
        for row in month_rows:
            inspector_name = row.get("inspector_name", "").strip()
            date_requested = row.get("date_requested", "").strip()
            time_requested = row.get("time_requested", "").strip()
            comment = row.get("comment", "").strip()
            truck_model_year = row.get("truck_model_year", "").strip()
            truck_unit_number = row.get("truck_unit_number", "").strip()
            items = parse_items_json(row.get("items_json", ""))

            ws1.cell(row=row_cursor, column=1, value="Date Requested")
            ws1.cell(row=row_cursor, column=1).font = bold
            ws1.cell(row=row_cursor, column=3, value=date_requested)

            ws1.cell(row=row_cursor, column=5, value="INSPECTOR")
            ws1.cell(row=row_cursor, column=5).font = bold
            ws1.cell(row=row_cursor, column=7, value=inspector_name)

            row_cursor += 1
            ws1.cell(row=row_cursor, column=1, value="Time Requested")
            ws1.cell(row=row_cursor, column=1).font = bold
            ws1.cell(row=row_cursor, column=3, value=time_requested)

            row_cursor += 2
            ws1.cell(row=row_cursor, column=1, value="QUANTITY")
            ws1.cell(row=row_cursor, column=1).font = bold
            ws1.cell(row=row_cursor, column=3, value="EQUIPMENT ITEM")
            ws1.cell(row=row_cursor, column=3).font = bold

            ws1.cell(row=row_cursor, column=1).border = bottom_border
            ws1.cell(row=row_cursor, column=3).border = bottom_border

            row_cursor += 2

            if items:
                for entry in items:
                    quantity = int(entry.get("quantity", 0) or 0)
                    item_name = str(entry.get("item", "")).strip()
                    value = entry.get("value", "")

                    item_text = item_name
                    if value_is_meaningful(value):
                        item_text = f"{item_text} ({value})"

                    ws1.cell(row=row_cursor, column=1, value=quantity)
                    ws1.cell(row=row_cursor, column=3, value=item_text)
                    row_cursor += 1
            else:
                ws1.cell(row=row_cursor, column=3, value="No items selected.")
                row_cursor += 1

            if truck_model_year or truck_unit_number:
                row_cursor += 1
                ws1.cell(row=row_cursor, column=1, value="DETAILS")
                ws1.cell(row=row_cursor, column=1).font = bold
                row_cursor += 1

                if truck_model_year:
                    ws1.cell(row=row_cursor, column=3, value=f"Truck Model Year: {truck_model_year}")
                    row_cursor += 1
                if truck_unit_number:
                    ws1.cell(row=row_cursor, column=3, value=f"Truck Unit Number: {truck_unit_number}")
                    row_cursor += 1

            if comment:
                row_cursor += 1
                ws1.cell(row=row_cursor, column=1, value="COMMENTS")
                ws1.cell(row=row_cursor, column=1).font = bold
                row_cursor += 1
                ws1.cell(row=row_cursor, column=3, value=comment)
                ws1.cell(row=row_cursor, column=3).alignment = Alignment(wrap_text=True, vertical="top")
                row_cursor += 1

            row_cursor += 3

    ws2 = wb.create_sheet(title="Monthly Item Totals")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 18

    ws2["A1"] = "EQUIPMENT ITEM"
    ws2["B1"] = "TOTAL REQUESTED"
    ws2["A1"].font = bold
    ws2["B1"].font = bold

    monthly_totals = compute_item_totals_from_rows(month_rows)
    output_row = 2
    for item_name, total_qty in monthly_totals.items():
        ws2.cell(row=output_row, column=1, value=item_name)
        ws2.cell(row=output_row, column=2, value=total_qty)
        output_row += 1

    if output_row == 2:
        ws2["A2"] = "No items found for selected month."

    file_buffer = io.BytesIO()
    wb.save(file_buffer)
    return file_buffer.getvalue()


# ---------- Admin UI ----------
def render_admin_manage_items(items: List[Item]) -> None:
    st.subheader("Add item")
    with st.form("admin_add_item", clear_on_submit=True):
        new_label = st.text_input("Item name", placeholder="Example: EXTRA BATTERY PACK")
        value_field = st.selectbox(
            "Optional right-side field",
            options=["none", "text", "number", "choice"],
            help="none = quantity button only. text/number = quantity button plus input. choice = quantity button plus dropdown.",
        )

        choice_text = ""
        if value_field == "choice":
            choice_text = st.text_input(
                "Choices (comma-separated)",
                value="S, M, L, XL",
                help="Example: S, M, L, XL, XXL",
            )

        add_item = st.form_submit_button("Add item to bottom", use_container_width=True)

        if add_item:
            if not new_label.strip():
                st.error("Item name is required.")
            else:
                new_item = Item(label=new_label.strip(), value_field=value_field)
                if value_field == "choice":
                    choices = [c.strip() for c in choice_text.split(",") if c.strip()]
                    new_item.choices = choices if choices else ["Option 1", "Option 2"]

                items.append(new_item)
                save_items(items)
                st.success("Added.")
                st.rerun()

    st.divider()
    st.subheader("Remove existing items")

    if items:
        for idx, item in enumerate(items):
            col_label, col_btn = st.columns([5, 1], vertical_alignment="center")
            with col_label:
                item_text = item.label
                if item.value_field == "choice" and item.choices:
                    item_text += f" ({', '.join(item.choices)})"
                elif item.value_field in {"text", "number"}:
                    item_text += f" ({item.value_field})"
                st.write(item_text)

            with col_btn:
                if st.button("Remove", key=f"remove_item_{idx}", use_container_width=True):
                    updated_items = load_items()
                    if 0 <= idx < len(updated_items):
                        updated_items.pop(idx)
                        save_items(updated_items)
                        st.success("Item removed.")
                        st.rerun()
    else:
        st.caption("No items to remove.")


def render_admin_popup_settings(admin_settings: Dict[str, Any]) -> None:
    st.subheader("Submit popup message")

    with st.form("admin_submit_popup_form"):
        popup_enabled = st.toggle(
            "Show custom popup after submit",
            value=admin_settings.get("submit_popup_enabled", False),
        )
        popup_message = st.text_area(
            "Custom popup message",
            value=admin_settings.get("submit_popup_message", ""),
            height=120,
            placeholder="Type the custom message users should see after they click Submit",
        )

        save_popup = st.form_submit_button("Save popup settings", use_container_width=True)

        if save_popup:
            save_admin_settings(
                {
                    "submit_popup_enabled": popup_enabled,
                    "submit_popup_message": popup_message,
                }
            )
            st.success("Popup settings saved.")
            st.rerun()


def render_admin_reports(items: List[Item]) -> None:
    history_rows = load_request_history_rows()
    all_months = get_available_months(history_rows)

    st.subheader("Monthly reports")

    if all_months:
        default_month = all_months[0]
        selected_month = st.selectbox("Choose month", options=all_months, index=0)
    else:
        default_month = datetime.now().strftime("%Y-%m")
        selected_month = st.selectbox("Choose month", options=[default_month], index=0)

    month_rows = [row for row in history_rows if row.get("year_month", "").strip() == selected_month]

    workbook_bytes = build_monthly_workbook(selected_month, month_rows)
    st.download_button(
        label=f"Download {selected_month} Excel workbook",
        data=workbook_bytes,
        file_name=f"Inspector_Request_Report_{selected_month}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    st.divider()
    st.subheader("Current item totals")

    month_totals = compute_item_totals_from_rows(month_rows)
    lifetime_totals = compute_lifetime_item_totals(history_rows)

    if items:
        for idx, item in enumerate(items):
            monthly_qty = month_totals.get(item.label, 0)
            lifetime_qty = lifetime_totals.get(item.label, 0)

            c1, c2, c3 = st.columns([5, 2, 2], vertical_alignment="center")
            with c1:
                st.markdown(f"**{item.label}**")
            with c2:
                st.metric(f"{selected_month}", monthly_qty)
            with c3:
                st.metric("Lifetime", lifetime_qty)
    else:
        st.caption("No active items found.")

    retired_items = [name for name in lifetime_totals.keys() if name not in {i.label for i in items}]
    if retired_items:
        st.divider()
        st.subheader("Removed items with saved history")
        for item_name in retired_items:
            c1, c2, c3 = st.columns([5, 2, 2], vertical_alignment="center")
            with c1:
                st.markdown(f"**{item_name}**")
            with c2:
                st.metric(f"{selected_month}", month_totals.get(item_name, 0))
            with c3:
                st.metric("Lifetime", lifetime_totals.get(item_name, 0))

    st.divider()
    st.subheader("Request history preview")

    if month_rows:
        preview_rows = []
        for row in reversed(month_rows[-25:]):
            preview_rows.append(
                {
                    "Name": row.get("inspector_name", ""),
                    "Date": row.get("date_requested", ""),
                    "Time": row.get("time_requested", ""),
                    "Items": row.get("items_display", ""),
                }
            )
        st.dataframe(preview_rows, use_container_width=True, hide_index=True)
    else:
        st.caption("No requests found for selected month.")

    st.divider()
    st.subheader("Raw CSV downloads")
    download_request_history_csv_button()
    download_csv_button()


# ---------- Public / main request form ----------
def init_form_session_state(items: List[Item]) -> None:
    for idx, item in enumerate(items):
        qty_key = f"qty_{idx}"
        if qty_key not in st.session_state:
            st.session_state[qty_key] = 0

        if item.value_field == "text":
            val_key = f"val_{idx}"
            if val_key not in st.session_state:
                st.session_state[val_key] = ""
        elif item.value_field == "number":
            val_key = f"val_{idx}"
            if val_key not in st.session_state:
                st.session_state[val_key] = 0
        elif item.value_field == "choice":
            val_key = f"val_{idx}"
            default_choice = item.choices[0] if item.choices else "Option 1"
            if val_key not in st.session_state:
                st.session_state[val_key] = default_choice


def clear_form_state(items: List[Item]) -> None:
    for idx, item in enumerate(items):
        st.session_state[f"qty_{idx}"] = 0
        if item.value_field == "text":
            st.session_state[f"val_{idx}"] = ""
        elif item.value_field == "number":
            st.session_state[f"val_{idx}"] = 0
        elif item.value_field == "choice":
            st.session_state[f"val_{idx}"] = item.choices[0] if item.choices else "Option 1"

    st.session_state["inspector_name_input"] = ""
    st.session_state["comment_input"] = ""


def build_submission_payload(items: List[Item]) -> Tuple[List[Dict[str, Any]], Optional[int], Optional[int]]:
    item_entries: List[Dict[str, Any]] = []
    truck_model_year_value: Optional[int] = None
    truck_unit_number_value: Optional[int] = None

    for idx, item in enumerate(items):
        qty = int(st.session_state.get(f"qty_{idx}", 0) or 0)
        value = st.session_state.get(f"val_{idx}") if item.value_field != "none" else None

        if is_truck_field(item.label):
            if item.label.strip().lower() == "truck model year" and value_is_meaningful(value):
                truck_model_year_value = int(value)
            if item.label.strip().lower() == "truck unit number" and value_is_meaningful(value):
                truck_unit_number_value = int(value)
            continue

        if qty > 0:
            item_entries.append(
                {
                    "item": item.label,
                    "quantity": qty,
                    "value": value if value_is_meaningful(value) else "",
                }
            )

    return item_entries, truck_model_year_value, truck_unit_number_value


def render_request_form(items: List[Item], admin_settings: Dict[str, Any]) -> None:
    st.subheader("Request form")

    init_form_session_state(items)

    inspector_name = st.text_input(
        "Inspector Name (required)",
        key="inspector_name_input",
        placeholder="Type inspector name",
    )

    st.markdown(
        "Tap **+1 NEED** to increase quantity for any equipment needed. "
        "At the end click submit. Then a download pdf button will appear. "
        "Download the pdf and email it."
    )
    st.divider()

    for idx, item in enumerate(items):
        qty_key = f"qty_{idx}"
        val_key = f"val_{idx}"

        if is_truck_field(item.label):
            st.markdown(f"**{item.label}**")
            if item.value_field == "number":
                st.number_input(
                    "",
                    key=val_key,
                    label_visibility="collapsed",
                    step=1,
                    format="%d",
                )
            elif item.value_field == "text":
                st.text_input(
                    "",
                    key=val_key,
                    label_visibility="collapsed",
                    placeholder="Enter value",
                )
            st.divider()
            continue

        c1, c2, c3 = st.columns([1.3, 0.8, 3.4], vertical_alignment="center")

        with c1:
            if st.button("+1 NEED", key=f"add_qty_{idx}", use_container_width=True, type="primary"):
                st.session_state[qty_key] = int(st.session_state.get(qty_key, 0) or 0) + 1
                st.rerun()

        with c2:
            st.markdown(f"**Qty: {int(st.session_state.get(qty_key, 0) or 0)}**")
            if st.button("Reset", key=f"reset_qty_{idx}", use_container_width=True):
                st.session_state[qty_key] = 0
                st.rerun()

        with c3:
            st.markdown(f"**{item.label}**")
            if item.value_field == "text":
                st.text_input(
                    "",
                    key=val_key,
                    label_visibility="collapsed",
                    placeholder="Enter text",
                )
            elif item.value_field == "number":
                st.number_input(
                    "",
                    key=val_key,
                    label_visibility="collapsed",
                    step=1,
                    format="%d",
                )
            elif item.value_field == "choice":
                choices = item.choices or ["Option 1", "Option 2"]
                current_val = st.session_state.get(val_key, choices[0])
                if current_val not in choices:
                    st.session_state[val_key] = choices[0]
                st.selectbox(
                    "",
                    options=choices,
                    key=val_key,
                    label_visibility="collapsed",
                )

        st.divider()

    comment = st.text_area(
        "Comments (optional)",
        key="comment_input",
        height=120,
        placeholder="Type any notes here...",
    )

    if "submitted" not in st.session_state:
        st.session_state.submitted = False
    if "last_pdf_bytes" not in st.session_state:
        st.session_state.last_pdf_bytes = None
    if "last_pdf_filename" not in st.session_state:
        st.session_state.last_pdf_filename = None
    if "last_success" not in st.session_state:
        st.session_state.last_success = False
    if "show_submit_popup" not in st.session_state:
        st.session_state.show_submit_popup = False
    if "submit_popup_message" not in st.session_state:
        st.session_state.submit_popup_message = ""

    submit = st.button(
        "Submit",
        type="primary",
        use_container_width=True,
        disabled=st.session_state.submitted,
    )

    if submit:
        st.session_state.submitted = True
        st.session_state.last_success = False
        st.session_state.last_pdf_bytes = None
        st.session_state.last_pdf_filename = None
        st.session_state.show_submit_popup = False
        st.session_state.submit_popup_message = ""

        st.warning("Do not refresh page. Wait for download button to appear.")

        if not inspector_name.strip():
            st.error("Inspector Name is required.")
            st.session_state.submitted = False
            st.stop()

        item_entries, truck_model_year_value, truck_unit_number_value = build_submission_payload(items)
        clean_comment = comment.strip()

        try:
            legacy_needed: List[Dict[str, Any]] = []
            for entry in item_entries:
                item_text = entry["item"]
                value = entry.get("value", "")
                qty = int(entry.get("quantity", 0) or 0)

                if value_is_meaningful(value):
                    legacy_needed.append({"item": item_text, "value": f"Qty {qty} | {value}"})
                else:
                    legacy_needed.append({"item": item_text, "value": f"Qty {qty}"})

            if truck_model_year_value not in (None, 0):
                legacy_needed.append({"item": "TRUCK MODEL YEAR", "value": int(truck_model_year_value)})

            if truck_unit_number_value not in (None, 0):
                legacy_needed.append({"item": "TRUCK UNIT NUMBER", "value": int(truck_unit_number_value)})

            pdf_filename, pdf_bytes = create_receipt_pdf(
                inspector_name=inspector_name,
                item_entries=item_entries,
                comment=clean_comment,
                truck_model_year=truck_model_year_value,
                truck_unit_number=truck_unit_number_value,
            )

            append_submission(inspector_name, legacy_needed, clean_comment)
            append_request_history(
                inspector_name=inspector_name,
                item_entries=item_entries,
                comment=clean_comment,
                truck_model_year=truck_model_year_value,
                truck_unit_number=truck_unit_number_value,
                pdf_filename=pdf_filename,
            )

            st.session_state.last_pdf_filename = pdf_filename
            st.session_state.last_pdf_bytes = pdf_bytes
            st.session_state.last_success = True

            latest_admin_settings = load_admin_settings()
            popup_enabled = latest_admin_settings.get("submit_popup_enabled", False)
            popup_message = latest_admin_settings.get("submit_popup_message", "").strip()

            if popup_enabled and popup_message:
                st.session_state.show_submit_popup = True
                st.session_state.submit_popup_message = popup_message

            clear_form_state(items)

        except Exception as e:
            st.error(f"Submit failed: {e}")
            st.session_state.submitted = False
            st.stop()

        st.session_state.submitted = False

    if (
        st.session_state.last_success
        and st.session_state.last_pdf_bytes
        and st.session_state.last_pdf_filename
    ):
        st.download_button(
            label="Download PDF Receipt",
            data=st.session_state.last_pdf_bytes,
            file_name=st.session_state.last_pdf_filename,
            mime="application/pdf",
            use_container_width=True,
        )

    if st.session_state.show_submit_popup and st.session_state.submit_popup_message:
        @st.dialog("Message")
        def show_custom_submit_popup() -> None:
            st.write(st.session_state.submit_popup_message)
            if st.button("Close", use_container_width=True):
                st.session_state.show_submit_popup = False
                st.rerun()

        show_custom_submit_popup()


# ---------- Main ----------
def main() -> None:
    st.set_page_config(page_title=APP_TITLE, layout="centered")

    st.markdown(
        """
        <style>
        section[data-testid="stSidebar"] button[kind="secondary"],
        div.stButton > button {
            min-height: 52px !important;
            font-size: 16px !important;
            font-weight: 700 !important;
            border-radius: 10px !important;
        }

        div.stButton > button {
            white-space: nowrap !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.title(APP_TITLE)

    items = load_items()
    admin_settings = load_admin_settings()

    params = st.query_params
    is_admin = str(params.get("admin", "0")).strip() == "1"

    if not is_admin:
        st.markdown(
            """
            <style>
              section[data-testid="stSidebar"] {display:none !important;}
              div[data-testid="collapsedControl"] {display:none !important;}
              header button {display:none !important;}
              .block-container {padding-left: 1rem !important; padding-right: 1rem !important;}
            </style>
            """,
            unsafe_allow_html=True,
        )

    if is_admin:
        st.caption("Admin mode enabled")
        tab_request, tab_items, tab_popup, tab_reports = st.tabs(
            [
                "Request Form",
                "Manage Items",
                "Popup Message",
                "Totals / History",
            ]
        )

        with tab_request:
            render_request_form(items, admin_settings)

        with tab_items:
            render_admin_manage_items(items)

        with tab_popup:
            render_admin_popup_settings(admin_settings)

        with tab_reports:
            render_admin_reports(items)

    else:
        render_request_form(items, admin_settings)


if __name__ == "__main__":
    main()
